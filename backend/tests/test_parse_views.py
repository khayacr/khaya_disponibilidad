"""
Tests de parsing de vistas/ubicación desde XLSX (fila 3 vs fila 4, merges).
"""
import os
import tempfile
from pathlib import Path

import openpyxl
import pytest

BACKEND = Path(__file__).resolve().parent.parent
import sys

sys.path.insert(0, str(BACKEND))

from server import (  # noqa: E402
    parse_xlsx_to_units,
    _should_use_detail_row_for_views,
    _detect_vista_row,
)


@pytest.fixture
def sheet_path_horizontal_merge_row3():
    """
    Torre E típica: B3:K3 combinado con un título; fila 4 con vista por columna.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Torre E"
    ws.merge_cells("B3:K3")
    ws["B3"] = "Vista Este"
    # Columnas B–K = aptos 1–10: alternar Este / Oeste
    for apt in range(1, 11):
        col = apt + 1
        ws.cell(4, col, value="Vista Oeste" if apt % 2 == 0 else "Vista Este Esquina")
    ws.cell(5, 1, value="Piso 14")
    for apt in range(1, 11):
        ws.cell(5, apt + 1, value=100000.0)
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    yield path
    os.unlink(path)


def test_horizontal_merge_row3_uses_row4_views(sheet_path_horizontal_merge_row3):
    units, towers = parse_xlsx_to_units(sheet_path_horizontal_merge_row3)
    assert len(towers) == 1
    u14 = [u for u in units if u["floor"] == 14]
    assert len(u14) == 10
    by_apt = {u["apartment"]: u["view"] for u in u14}
    assert by_apt[1] == "Vista Este Esquina"
    assert by_apt[2] == "Vista Oeste"
    assert by_apt[3] == "Vista Este Esquina"


def test_should_use_detail_row_detects_merge():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Torre E"
    ws.merge_cells("B3:K3")
    ws["B3"] = "X"
    assert _should_use_detail_row_for_views(ws) is True


def test_should_use_detail_row_all_ten_same_text():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Torre E"
    for c in range(2, 12):
        ws.cell(3, c, value="Igual")
    assert _should_use_detail_row_for_views(ws) is True


@pytest.fixture
def sheet_path_torre_e_apto_row_layout():
    """Misma estructura que la hoja real: fila 3 Apto N, fila 4 vistas, fila 5 Piso 14."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Torre E"
    for apt in range(1, 11):
        ws.cell(3, apt + 1, value=f"Apto {apt}")
    for apt in range(1, 11):
        ws.cell(4, apt + 1, value="Vista Oeste" if apt % 2 == 0 else "Vista Este Esquina")
    ws.cell(5, 1, value="Piso 14")
    for apt in range(1, 11):
        ws.cell(5, apt + 1, value=100000.0)
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    yield path
    os.unlink(path)


def test_detect_vista_row_apto_in_row3(sheet_path_torre_e_apto_row_layout):
    wb = openpyxl.load_workbook(sheet_path_torre_e_apto_row_layout, data_only=True)
    ws = wb.active
    assert _detect_vista_row(ws) == 4
    wb.close()


def test_apto_row_layout_reads_row4_views(sheet_path_torre_e_apto_row_layout):
    units, _ = parse_xlsx_to_units(sheet_path_torre_e_apto_row_layout)
    u14 = [u for u in units if u["floor"] == 14]
    by_apt = {u["apartment"]: u["view"] for u in u14}
    assert by_apt[1] == "Vista Este Esquina"
    assert by_apt[2] == "Vista Oeste"
