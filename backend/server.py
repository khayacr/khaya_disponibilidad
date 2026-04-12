from fastapi import FastAPI, APIRouter, HTTPException
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import os
import sys
import json
import logging
import httpx
import tempfile
import asyncio
from contextlib import asynccontextmanager
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Tuple
import uuid
import re
from datetime import datetime, timezone
import openpyxl
from openpyxl.utils import range_boundaries

# Google Sheets write-back (gspread + google-auth, already in requirements.txt)
GSPREAD_IMPORT_ERROR: Optional[str] = None
try:
    import gspread
    from google.oauth2 import service_account
    GSPREAD_AVAILABLE = True
except ImportError as e:
    GSPREAD_AVAILABLE = False
    GSPREAD_IMPORT_ERROR = str(e)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# ID de la hoja (extrae del enlace: docs.google.com/spreadsheets/d/<ID>/…)
SPREADSHEET_ID = os.environ.get("SPREADSHEET_ID", "15XS6yetq6XgS-wpMxpu5exUNUsXFH3UstSac7HvTblY")
XLSX_DOWNLOAD_URL = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"
# Enlace de edición (solo informativo; la app lee vía export XLSX sin API ni credenciales)
GOOGLE_SHEETS_EDITOR_URL = os.environ.get(
    "GOOGLE_SHEETS_EDITOR_URL",
    f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/edit",
)
SYNC_INTERVAL_SECONDS = 60
GOOGLE_CREDENTIALS_JSON = os.environ.get('GOOGLE_CREDENTIALS_JSON', '')

# ── IN-MEMORY STORAGE (no MongoDB needed) ─────────────────────────
UNITS: List[dict] = []
TOWERS: dict = {}
STATUS_CHECKS_LIST: List[dict] = []
CONTACT_REQUESTS_LIST: List[dict] = []

last_sync_time: Optional[str] = None
sync_task: Optional[asyncio.Task] = None
units_initialized: bool = False

api_router = APIRouter(prefix="/api")

# ============== MODELS ==============

class StatusCheck(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_name: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StatusCheckCreate(BaseModel):
    client_name: str

class UnitInfo(BaseModel):
    id: str
    code: str
    floor: int
    apartment: int
    price: float

class ContactRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    email: EmailStr
    phone: str
    message: Optional[str] = None
    unit: Optional[UnitInfo] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    status: str = "pending"

class ContactRequestCreate(BaseModel):
    name: str
    email: EmailStr
    phone: str
    message: Optional[str] = None
    unit: Optional[UnitInfo] = None

class Unit(BaseModel):
    id: str
    code: str
    tower: str
    floor: int
    apartment: int
    price: float
    view: str
    viewDirection: str
    type: str
    status: str
    area: float
    parkingArea: float
    totalArea: float
    bedrooms: int
    bathrooms: int
    delivery: str
    rentability: float
    ubicacion: str = ""

class UnitUpdate(BaseModel):
    price: Optional[float] = None
    status: Optional[str] = None
    area: Optional[float] = None
    parkingArea: Optional[float] = None
    bedrooms: Optional[int] = None
    bathrooms: Optional[int] = None
    tower: Optional[str] = None


class BulkFloorStatusUpdate(BaseModel):
    """Cambiar el estado de las 10 unidades de un piso (y sincronizar con Sheets en serie)."""
    tower: str
    floor: int
    status: str

class Tower(BaseModel):
    name: str
    floors: int
    apartments: int
    delivery: str

# ============== GOOGLE SHEETS SYNC (READ) ==============

def get_cell_color_rgb(cell) -> str:
    """Extract hex RGB color string from a cell's fill, handling theme colors."""
    fill = cell.fill
    if not fill or not fill.start_color:
        return "NONE"
    color = fill.start_color
    if color.type == 'rgb':
        try:
            rgb = color.rgb
            if isinstance(rgb, str) and len(rgb) >= 6:
                return rgb.upper()
        except Exception:
            pass
    return "THEME"

def color_to_status(rgb: str) -> str:
    """Map Excel cell background color to unit status."""
    if not rgb or rgb in ("NONE", "THEME", "00000000"):
        return "Disponible"
    rgb_upper = rgb.upper()
    if "FF0000" in rgb_upper:
        return "Vendido"
    if rgb_upper in ("FFA5A5A5",):
        return "Bloqueado"
    if rgb_upper in ("FF83CAEB",):
        return "Reservado"
    if "B8D4EF" in rgb_upper:
        return "Disponible"
    if "F09200" in rgb_upper:
        return "Disponible"
    return "Disponible"

def _download_xlsx_via_service_account() -> str:
    """
    Export XLSX usando la cuenta de servicio (misma que gspread).
    Necesario cuando la hoja NO está en «Cualquiera con el enlace» y el GET público devuelve 401.
    """
    from google.auth.transport.requests import AuthorizedSession

    creds_dict = json.loads(GOOGLE_CREDENTIALS_JSON)
    creds = service_account.Credentials.from_service_account_info(
        creds_dict,
        scopes=["https://www.googleapis.com/auth/spreadsheets"],
    )
    session = AuthorizedSession(creds)
    r = session.get(XLSX_DOWNLOAD_URL, timeout=120)
    if r.status_code != 200:
        raise Exception(f"Authenticated XLSX download failed: HTTP {r.status_code}")
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(r.content)
    tmp.close()
    return tmp.name


async def download_xlsx() -> str:
    """Download the XLSX file from Google Sheets export URL; fallback autenticado si la hoja es privada."""
    async with httpx.AsyncClient(timeout=60.0) as http_client:
        response = await http_client.get(XLSX_DOWNLOAD_URL, follow_redirects=True)

    if response.status_code == 200:
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.write(response.content)
        tmp.close()
        return tmp.name

    if response.status_code in (401, 403) and GOOGLE_CREDENTIALS_JSON and GSPREAD_AVAILABLE:
        logging.info(
            "Public export HTTP %s — usando descarga autenticada (hoja no pública o restricción de enlace).",
            response.status_code,
        )
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, _download_xlsx_via_service_account)

    raise Exception(f"Failed to download: HTTP {response.status_code}")

def _merge_bounds_for_cell(ws, row: int, col: int):
    """Si (row,col) está en un rango combinado, devuelve (min_col, min_row, max_col, max_row)."""
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return min_col, min_row, max_col, max_row
    return None


def _row3_is_horizontal_banner_cell(ws, col: int) -> bool:
    """
    True si la celda (fila 3, col) forma parte de un merge horizontal en la fila 3
    (varias columnas, una sola fila). En ese caso openpyxl repite el mismo texto en
    todas las columnas y no hay vista por apartamento en la fila 3.
    """
    b = _merge_bounds_for_cell(ws, 3, col)
    if not b:
        return False
    min_col, min_row, max_col, max_row = b
    if min_row != max_row or min_row != 3:
        return False
    return (max_col - min_col + 1) > 1


def _effective_cell_value(ws, row: int, col: int):
    """
    Valor legible de una celda. Si pertenece a un rango combinado (merge),
    Excel/openpyxl solo guarda el valor en la esquina superior izquierda;
    el resto llega como None y antes caíamos siempre en 'Vista Este'.
    """
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return ws.cell(min_row, min_col).value
    return ws.cell(row, col).value

def _vista_raw_for_apartment_column(ws, col: int, primary_row: int, prefer_detail_row: bool):
    """
    Texto de vista por columna (B–K). primary_row es 3 (F/G) o 4 (E con fila «Apto»).
    Si prefer_detail_row=True, la fila 3 es banner combinado: no usarla; prueba 4+.
    Si primary_row=4, no se usa la fila 3 (etiquetas Apto), solo 4,5,6,2.
    """
    if prefer_detail_row:
        for r in (4, 5, 6, 2):
            raw = _effective_cell_value(ws, r, col)
            if _looks_like_vista_text(raw):
                return raw
        return None
    if primary_row == 4:
        for r in (4, 5, 6, 2):
            raw = _effective_cell_value(ws, r, col)
            if _looks_like_vista_text(raw):
                return raw
        return None
    raw = _effective_cell_value(ws, 3, col)
    if _looks_like_vista_text(raw):
        return raw
    for r in (4, 5, 6, 2):
        raw = _effective_cell_value(ws, r, col)
        if _looks_like_vista_text(raw):
            return raw
    return raw


def _looks_like_vista_text(raw) -> bool:
    """Descarta precios numéricos cuando se busca vista en filas 4+ (p. ej. banner combinado en F/G)."""
    if raw is None:
        return False
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return False
    s = str(raw).strip()
    if not s:
        return False
    low = s.replace(".", "").replace(",", "").replace(" ", "")
    if low.isdigit():
        return False
    return True


def _detect_vista_row(ws) -> int:
    """
    Torre E tiene una fila extra: fila 3 = «Apto 1»…«Apto 10», fila 4 = vistas.
    En F/G la fila 3 es ya la de vistas. Si B3 empieza por «Apto», leer vistas en fila 4.
    """
    raw = _effective_cell_value(ws, 3, 2)
    s = str(raw).strip().lower() if raw is not None else ""
    if s.startswith("apto"):
        return 4
    return 3


def _should_use_detail_row_for_views(ws) -> bool:
    """
    Usar filas 4+ para vistas por columna cuando la fila 3 no distingue apartamentos:
    - merge horizontal B:K en fila 3, o
    - el mismo texto no vacío en columnas 2–11 (típico de banner combinado).
    """
    if any(_row3_is_horizontal_banner_cell(ws, c) for c in range(2, 12)):
        return True
    r3 = []
    for c in range(2, 12):
        v = _effective_cell_value(ws, 3, c)
        r3.append(str(v).strip() if v is not None else "")
    non_empty = [x for x in r3 if x]
    # Las 10 columnas con el mismo texto = título combinado, no vista por apartamento.
    if len(non_empty) == 10 and len(set(non_empty)) == 1:
        return True
    return False

def _parse_vista_ubicacion_row3_cell(raw) -> Tuple[str, str]:
    """
    Fila 3 de cada hoja: en cada columna de apartamento puede venir vista y ubicación.
    - Dos líneas en la misma celda: 1ª = texto para vista (Este/Oeste/Esquina…), 2ª = ubicación.
    - Una línea: si hay separador (· | / — -), izquierda = vista, derecha = ubicación.
    """
    if raw is None:
        return "Vista Este", ""
    text = str(raw).strip()
    if not text:
        return "Vista Este", ""
    lines = [ln.strip() for ln in re.split(r"\r?\n+", text) if ln.strip()]
    if len(lines) >= 2:
        return lines[0], " · ".join(lines[1:])
    for sep in (" · ", " | ", " / ", " — ", " – ", " - "):
        if sep in text:
            a, b = text.split(sep, 1)
            a, b = a.strip(), b.strip()
            if b:
                return a, b
    return text, ""

def _normalize_view_label_and_meta(vista_line: str) -> Tuple[str, str, str]:
    """
    A partir del texto de vista (1ª línea, fila 3) devuelve:
    - view: una de Vista Este | Vista Este Esquina | Vista Oeste | Vista Oeste Esquina
    - viewDirection: Este | Oeste
    - type: Central | Esquinero

    Importante: en Python "este" in "oeste" es True (substring). Por eso se evalúa
    Oeste antes que Este.
    """
    t = (vista_line or "").strip()
    if not t:
        return "Vista Este", "Este", "Central"
    low = t.lower()
    has_esq = "esquina" in low
    if "oeste" in low:
        view = "Vista Oeste Esquina" if has_esq else "Vista Oeste"
        return view, "Oeste", "Esquinero" if has_esq else "Central"
    if "este" in low:
        view = "Vista Este Esquina" if has_esq else "Vista Este"
        return view, "Este", "Esquinero" if has_esq else "Central"
    return "Vista Este", "Este", "Central"

def parse_xlsx_to_units(filepath: str) -> tuple:
    """Parse the downloaded XLSX file into units and towers."""
    wb_values = openpyxl.load_workbook(filepath, data_only=True)
    wb_colors = openpyxl.load_workbook(filepath, data_only=False)
    all_units = []
    towers = []

    for sheet_name in wb_values.sheetnames:
        if not sheet_name.startswith('Torre'):
            continue

        ws_val = wb_values[sheet_name]
        ws_col = wb_colors[sheet_name]
        tower_letter = sheet_name.replace('Torre ', '')

        # Vistas por columna B–K: fila 3 en F/G; Torre E tiene fila 3 «Apto N» y vistas en fila 4.
        detail_row = _should_use_detail_row_for_views(ws_val)
        vista_row = _detect_vista_row(ws_val)
        view_map: dict = {}
        for col in range(2, 12):
            raw = _vista_raw_for_apartment_column(ws_val, col, vista_row, detail_row)
            vista_part, ubicacion_str = _parse_vista_ubicacion_row3_cell(raw)
            apt_num = col - 1
            view_label, direction, unit_type = _normalize_view_label_and_meta(vista_part)
            view_map[apt_num] = {
                "view": view_label,
                "viewDirection": direction,
                "type": unit_type,
                "ubicacion": ubicacion_str.strip(),
            }

        delivery_raw = _effective_cell_value(ws_val, vista_row, 12)
        if delivery_raw is None or (isinstance(delivery_raw, str) and not str(delivery_raw).strip()):
            delivery_raw = _effective_cell_value(ws_val, vista_row + 1, 12)
        if delivery_raw is None or (isinstance(delivery_raw, str) and not str(delivery_raw).strip()):
            delivery_raw = _effective_cell_value(ws_val, 3, 12)
        if delivery_raw is None or (isinstance(delivery_raw, str) and not str(delivery_raw).strip()):
            delivery_raw = _effective_cell_value(ws_val, 4, 12)
        delivery = "Mayo 2027"
        if delivery_raw:
            if isinstance(delivery_raw, datetime):
                month_names = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                               'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
                delivery = f"{month_names[delivery_raw.month]} {delivery_raw.year}"
            elif isinstance(delivery_raw, str) and '/' in delivery_raw:
                try:
                    parts = delivery_raw.split('/')
                    month_names = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                                   'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
                    delivery = f"{month_names[int(parts[0])]} {parts[2]}"
                except Exception:
                    pass

        max_floor = 0
        for row_idx in range(4, ws_val.max_row + 1):
            floor_cell = ws_val.cell(row_idx, 1).value
            if not floor_cell or not str(floor_cell).startswith('Piso'):
                continue
            try:
                floor = int(str(floor_cell).replace('Piso ', ''))
            except ValueError:
                continue

            if floor > max_floor:
                max_floor = floor

            for apt in range(1, 11):
                col = apt + 1
                val_cell = ws_val.cell(row_idx, col)
                col_cell = ws_col.cell(row_idx, col)
                price = val_cell.value

                if price is None or not isinstance(price, (int, float)) or price <= 0:
                    continue

                rgb = get_cell_color_rgb(col_cell)
                status = color_to_status(rgb)

                info = view_map.get(
                    apt,
                    {
                        "view": "Vista Este",
                        "viewDirection": "Este",
                        "type": "Central",
                        "ubicacion": "",
                    },
                )
                is_esquinero = info['type'] == 'Esquinero'
                base_area = 67.0
                parking_area = 14.3
                unit_id = f"{tower_letter}-{floor}-{apt}"

                all_units.append({
                    "id": unit_id,
                    "code": f"{tower_letter}-{str(floor).zfill(2)}-{apt}",
                    "tower": sheet_name,
                    "floor": floor,
                    "apartment": apt,
                    "price": float(price),
                    "view": info['view'],
                    "viewDirection": info['viewDirection'],
                    "type": info['type'],
                    "ubicacion": info.get('ubicacion', ''),
                    "status": status,
                    "area": base_area,
                    "parkingArea": parking_area,
                    "totalArea": base_area + parking_area,
                    "bedrooms": 2,
                    "bathrooms": 2 if is_esquinero else 1,
                    "delivery": delivery,
                    "rentability": 8.0
                })

        towers.append({
            "name": sheet_name,
            "floors": max_floor,
            "apartments": 10,
            "delivery": delivery
        })

    wb_values.close()
    wb_colors.close()
    return all_units, towers

async def sync_from_google_sheets():
    """Download XLSX, parse units, apply manual overrides, update in-memory state."""
    global last_sync_time, UNITS, TOWERS, units_initialized

    logging.info("Starting sync from Google Sheets...")
    filepath = await download_xlsx()
    try:
        all_units, towers_list = parse_xlsx_to_units(filepath)
    finally:
        os.unlink(filepath)

    if not all_units:
        logging.error("No units parsed from sheet")
        return False

    UNITS = all_units
    TOWERS = {t["name"]: t for t in towers_list}
    if not TOWERS:
        TOWERS = {"Torre E": {"name": "Torre E", "floors": 14, "apartments": 10, "delivery": "Mayo 2027"}}

    last_sync_time = datetime.now(timezone.utc).isoformat()
    units_initialized = True
    logging.info(f"Synced {len(UNITS)} units across {len(TOWERS)} towers at {last_sync_time}")
    return True

async def periodic_sync():
    """Background task: sync from Google Sheets every SYNC_INTERVAL_SECONDS."""
    while True:
        await asyncio.sleep(SYNC_INTERVAL_SECONDS)
        try:
            await sync_from_google_sheets()
        except Exception as e:
            logging.error(f"Periodic sync failed: {e}")

# ============== GOOGLE SHEETS WRITE-BACK ==============

# Status → Google Sheets API background color (float 0-1)
STATUS_COLORS_FLOAT = {
    'Vendido':    {'red': 1.0,   'green': 0.0,   'blue': 0.0},
    'Bloqueado':  {'red': 0.647, 'green': 0.647, 'blue': 0.647},
    'Reservado':  {'red': 0.514, 'green': 0.792, 'blue': 0.922},
    'Disponible': {'red': 0.722, 'green': 0.831, 'blue': 0.937},
}

def _write_status_to_sheet_sync(unit_id: str, new_status: str) -> tuple:
    """Returns (success: bool, client_message: str). client_message vacío si OK; si falla, texto para mostrar al usuario."""
    if not GOOGLE_CREDENTIALS_JSON:
        logging.warning("GOOGLE_CREDENTIALS_JSON not set — skipping Sheet write-back")
        return False, ""
    if not GSPREAD_AVAILABLE:
        logging.warning("gspread not available — skipping Sheet write-back")
        return False, ""

    try:
        parts = unit_id.split('-')
        if len(parts) != 3:
            logging.warning(f"Unexpected unit_id format: {unit_id}")
            return False, ""
        tower_letter, floor_str, apt_str = parts
        floor = int(floor_str)
        apartment = int(apt_str)
        sheet_name = f"Torre {tower_letter}"
        col = apartment + 1  # apt 1 → col B (2), apt 10 → col K (11)

        color = STATUS_COLORS_FLOAT.get(new_status, STATUS_COLORS_FLOAT['Disponible'])

        creds_dict = json.loads(GOOGLE_CREDENTIALS_JSON)
        creds = service_account.Credentials.from_service_account_info(
            creds_dict,
            scopes=['https://www.googleapis.com/auth/spreadsheets']
        )
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(SPREADSHEET_ID)
        ws = sh.worksheet(sheet_name)

        # Find the row labeled "Piso {floor}" in column A
        col_a = ws.col_values(1)
        row = None
        for i, val in enumerate(col_a):
            if val and str(val).strip() == f"Piso {floor}":
                row = i + 1  # gspread is 1-indexed
                break

        if row is None:
            logging.warning(f"Row for Piso {floor} not found in {sheet_name}")
            return False, ""

        cell = gspread.utils.rowcol_to_a1(row, col)
        ws.format(cell, {"backgroundColor": color})
        logging.info(f"Sheet write-back: {sheet_name}!{cell} → {new_status}")
        return True, ""

    except Exception as e:
        logging.exception("Sheet write-back failed for %s", unit_id)
        err = str(e).lower()
        if "not supported for this document" in err:
            return False, (
                "Google no permite cambiar el color en ESTE archivo (suele pasar con hojas importadas desde Excel). "
                "Solución: en Google Sheets → Archivo → «Guardar como hoja de cálculo de Google», "
                "o Archivo → «Hacer una copia» (hoja nativa). Pon el nuevo ID de la hoja en SPREADSHEET_ID en Render "
                "y comparte la copia con el email de la cuenta de servicio (Editor)."
            )
        return False, ""

async def write_status_to_sheet(unit_id: str, new_status: str) -> tuple:
    """Async wrapper: returns (success, client_error_message)."""
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _write_status_to_sheet_sync, unit_id, new_status)

# ============== APP LIFECYCLE ==============

@asynccontextmanager
async def lifespan(app: FastAPI):
    global sync_task
    logging.info("Python ejecutando la API: %s", sys.executable)
    logging.info(
        "gspread disponible: %s%s",
        GSPREAD_AVAILABLE,
        f" — {GSPREAD_IMPORT_ERROR}" if GSPREAD_IMPORT_ERROR else "",
    )
    sync_task = asyncio.create_task(periodic_sync())
    logging.info(f"Started periodic sync task (every {SYNC_INTERVAL_SECONDS}s)")
    yield
    if sync_task:
        sync_task.cancel()
        try:
            await sync_task
        except asyncio.CancelledError:
            pass

app = FastAPI(title="KHAYA Real Estate API", lifespan=lifespan)

# ============== HELPERS ==============

async def get_units_from_db():
    """Returns in-memory units, triggering initial sync if needed."""
    if not units_initialized:
        await sync_from_google_sheets()
    return UNITS

# ============== ROUTES ==============

@api_router.get("/")
async def root():
    return {"message": "KHAYA Real Estate API", "version": "2.0.0"}

@api_router.get("/health")
async def health():
    return {"status": "healthy"}

@api_router.post("/sync")
async def sync_sheets():
    """Manually trigger sync from Google Sheets."""
    success = await sync_from_google_sheets()
    if success:
        return {
            "status": "success",
            "message": "Data synced from Google Sheets",
            "units": len(UNITS),
            "towers": list(TOWERS.keys()),
            "last_sync": last_sync_time
        }
    raise HTTPException(status_code=500, detail="Failed to sync from Google Sheets")

@api_router.get("/sync-status")
async def get_sync_status():
    return {
        "last_sync": last_sync_time,
        "interval_seconds": SYNC_INTERVAL_SECONDS,
        "units_count": len(UNITS)
    }


def _sheet_write_enabled() -> bool:
    return bool(GOOGLE_CREDENTIALS_JSON) and bool(GSPREAD_AVAILABLE)


def _sheet_write_disabled_detail() -> str:
    """Mensaje cuando el usuario intenta cambiar estado sin escritura a Sheets (credenciales o gspread)."""
    # Mensajes separados: antes mezclábamos todo y parecía "la IA no puede". Ver docs/ESCRITURA_GOOGLE_SHEETS.md
    if not GSPREAD_AVAILABLE:
        return (
            "Escritura desactivada: en este proceso no se pudo importar gspread. "
            "El proyecto ya lista gspread en backend/requirements.txt; en el servidor debe ejecutarse "
            "`pip install -r backend/requirements.txt` (revisa el build). En local: "
            "`python3 -m pip install gspread google-auth`. "
            "Guía: docs/ESCRITURA_GOOGLE_SHEETS.md — "
            f"Hoja: {GOOGLE_SHEETS_EDITOR_URL}"
        )
    if not GOOGLE_CREDENTIALS_JSON:
        return (
            "Falta GOOGLE_CREDENTIALS_JSON: cuenta de servicio en Google Cloud, JSON en la variable de entorno, "
            "y la hoja compartida con el client_email del JSON (Editor). "
            "Guía paso a paso: docs/ESCRITURA_GOOGLE_SHEETS.md — "
            f"Hoja: {GOOGLE_SHEETS_EDITOR_URL}"
        )
    return f"Escritura en la hoja no disponible. Hoja: {GOOGLE_SHEETS_EDITOR_URL}"


@api_router.get("/sheets-enabled")
async def sheets_enabled():
    """Lectura: export XLSX público (sin credenciales si la hoja es visible con el enlace). Escritura vía API solo si hay service account."""
    write_on = _sheet_write_enabled()
    return {
        "enabled": write_on,
        "write_enabled": write_on,
        "read_via_export": True,
        "mode": "api_writeback" if write_on else "editor_only",
        "spreadsheet_id": SPREADSHEET_ID,
        "spreadsheet_editor_url": GOOGLE_SHEETS_EDITOR_URL,
        "has_credentials": bool(GOOGLE_CREDENTIALS_JSON),
        "gspread_available": bool(GSPREAD_AVAILABLE),
        # Motivo concreto si falla el import (p. ej. "No module named 'gspread'") — útil en Render/local
        "gspread_import_error": GSPREAD_IMPORT_ERROR,
    }

@api_router.post("/status", response_model=StatusCheck)
async def create_status_check(input: StatusCheckCreate):
    obj = StatusCheck(**input.model_dump())
    STATUS_CHECKS_LIST.append(obj.model_dump())
    return obj

@api_router.get("/status", response_model=List[StatusCheck])
async def get_status_checks():
    return STATUS_CHECKS_LIST

@api_router.get("/towers")
async def get_towers():
    await get_units_from_db()
    return TOWERS

@api_router.get("/units", response_model=List[Unit])
async def get_units(
    tower: Optional[str] = None,
    floor: Optional[int] = None,
    view: Optional[str] = None,
    status: Optional[str] = None,
    unit_type: Optional[str] = None,
    min_price: Optional[float] = None,
    max_price: Optional[float] = None
):
    units = await get_units_from_db()
    filtered = units
    if tower is not None:
        filtered = [u for u in filtered if u["tower"] == tower]
    if floor is not None:
        filtered = [u for u in filtered if u["floor"] == floor]
    if view is not None:
        filtered = [u for u in filtered if u["view"] == view]
    if status is not None:
        filtered = [u for u in filtered if u["status"] == status]
    if unit_type is not None:
        filtered = [u for u in filtered if u["type"] == unit_type]
    if min_price is not None:
        filtered = [u for u in filtered if u["price"] >= min_price]
    if max_price is not None:
        filtered = [u for u in filtered if u["price"] <= max_price]
    return filtered

@api_router.get("/units/{unit_id}", response_model=Unit)
async def get_unit(unit_id: str):
    units = await get_units_from_db()
    for unit in units:
        if unit["id"] == unit_id:
            return unit
    raise HTTPException(status_code=404, detail="Unit not found")

@api_router.put("/units/{unit_id}", response_model=Unit)
async def update_unit(unit_id: str, update_data: UnitUpdate):
    # Misma carga inicial que GET /units: sin esto, UNITS puede estar vacío hasta el primer GET.
    await get_units_from_db()
    unit_id = unit_id.strip()
    unit = next((u for u in UNITS if u["id"] == unit_id), None)
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")

    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}

    if update_dict:
        if "area" in update_dict or "parkingArea" in update_dict:
            new_area = update_dict.get("area", unit["area"])
            new_parking = update_dict.get("parkingArea", unit["parkingArea"])
            update_dict["totalArea"] = new_area + new_parking

        if "status" in update_dict:
            if not _sheet_write_enabled():
                raise HTTPException(
                    status_code=503,
                    detail=_sheet_write_disabled_detail(),
                )
            # Write the new color back to Google Sheets (single source of truth)
            wrote, write_err = await write_status_to_sheet(unit_id, update_dict["status"])
            if not wrote:
                _parts = unit_id.split("-")
                _floor_hint = _parts[1] if len(_parts) >= 2 else "?"
                generic = (
                    "No se pudo escribir en Google Sheets. Revisa: JSON en GOOGLE_CREDENTIALS_JSON, "
                    "hoja compartida con client_email (Editor), SPREADSHEET_ID correcto, "
                    f"pestaña «Torre …» y fila «Piso {_floor_hint}». Ver logs del servidor."
                )
                raise HTTPException(
                    status_code=502,
                    detail=(write_err.strip() if write_err and write_err.strip() else generic),
                )
            # Re-sync to reflect what the sheet actually has
            await sync_from_google_sheets()
            updated = next((u for u in UNITS if u["id"] == unit_id), None)
            if not updated:
                raise HTTPException(
                    status_code=404,
                    detail=(
                        "Unidad no encontrada tras sincronizar: revisa que la celda tenga precio numérico "
                        f"válido en la hoja. id={unit_id}"
                    ),
                )
            return updated

        # Non-status edits remain in-memory only (sheet is treated as source of truth for status)
        unit.update(update_dict)

    return unit


VALID_STATUSES = frozenset({"Disponible", "Reservado", "Vendido", "Bloqueado"})


@api_router.put("/units/bulk-floor")
async def bulk_floor_status(body: BulkFloorStatusUpdate):
    """
    Actualiza el estado de todos los apartamentos (1–10) de un piso y escribe cada celda
    en Google Sheets de forma secuencial (evita condiciones de carrera y límites de API).
    """
    await get_units_from_db()
    if body.status not in VALID_STATUSES:
        raise HTTPException(status_code=400, detail=f"Estado no válido: {body.status}")
    if not _sheet_write_enabled():
        raise HTTPException(
            status_code=503,
            detail=_sheet_write_disabled_detail(),
        )
    tower_prefix = body.tower.replace("Torre ", "").strip()
    if not tower_prefix:
        raise HTTPException(status_code=400, detail="Torre no válida (ej. Torre E)")

    updated_ids: List[str] = []
    for apt in range(1, 11):
        unit_id = f"{tower_prefix}-{body.floor}-{apt}"
        unit = next((u for u in UNITS if u["id"] == unit_id), None)
        if not unit:
            raise HTTPException(status_code=404, detail=f"Unidad no encontrada: {unit_id}")
        updated_ids.append(unit_id)

    for uid in updated_ids:
        ok, msg = await write_status_to_sheet(uid, body.status)
        if not ok:
            raise HTTPException(
                status_code=502,
                detail=(
                    msg.strip()
                    if msg and msg.strip()
                    else f"No se pudo escribir en Google Sheets (unidad {uid}). Revisa credenciales, permisos y logs."
                ),
            )
        await asyncio.sleep(0.2)

    # Refresh from sheet so UI shows sheet truth
    await sync_from_google_sheets()

    return {
        "status": "success",
        "tower": body.tower,
        "floor": body.floor,
        "new_status": body.status,
        "count": len(updated_ids),
        "unit_ids": updated_ids,
    }


@api_router.get("/units/floor/{floor}", response_model=List[Unit])
async def get_units_by_floor(floor: int):
    units = await get_units_from_db()
    return [u for u in units if u["floor"] == floor]

@api_router.get("/stats")
async def get_stats():
    units = await get_units_from_db()
    total = len(units)
    disponibles = sum(1 for u in units if u["status"] == "Disponible")
    reservados  = sum(1 for u in units if u["status"] == "Reservado")
    vendidos    = sum(1 for u in units if u["status"] == "Vendido")
    bloqueados  = sum(1 for u in units if u["status"] == "Bloqueado")
    return {
        "total": total,
        "disponibles": disponibles,
        "reservados": reservados,
        "vendidos": vendidos,
        "bloqueados": bloqueados,
        "percentage": {
            "disponibles": round((disponibles / total) * 100) if total > 0 else 0,
            "reservados":  round((reservados  / total) * 100) if total > 0 else 0,
            "vendidos":    round((vendidos    / total) * 100) if total > 0 else 0,
            "bloqueados":  round((bloqueados  / total) * 100) if total > 0 else 0,
        }
    }

@api_router.get("/filters")
async def get_filters():
    units = await get_units_from_db()
    prices = [u["price"] for u in units]
    return {
        "towers":     sorted(set(u["tower"]  for u in units)),
        "floors":     sorted(set(u["floor"]  for u in units), reverse=True),
        "views":      sorted(set(u["view"]   for u in units)),
        "statuses":   sorted(set(u["status"] for u in units)),
        "types":      sorted(set(u["type"]   for u in units)),
        "priceRange": {
            "min": min(prices) if prices else 0,
            "max": max(prices) if prices else 0
        }
    }

@api_router.post("/contact", response_model=ContactRequest)
async def create_contact_request(input: ContactRequestCreate):
    obj = ContactRequest(**input.model_dump())
    doc = obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    CONTACT_REQUESTS_LIST.append(doc)
    return obj

@api_router.get("/contact", response_model=List[ContactRequest])
async def get_contact_requests():
    result = []
    for req in CONTACT_REQUESTS_LIST:
        r = dict(req)
        if isinstance(r.get('created_at'), str):
            r['created_at'] = datetime.fromisoformat(r['created_at'])
        result.append(r)
    return result

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve React static build (for Render single-service deployment)
from fastapi.staticfiles import StaticFiles
from starlette.responses import FileResponse

static_build_dir = ROOT_DIR / "static"
if static_build_dir.exists():
    react_assets_dir = static_build_dir / "static"
    if react_assets_dir.exists():
        app.mount("/static", StaticFiles(directory=str(react_assets_dir)), name="react-assets")

    @app.get("/{full_path:path}")
    async def serve_react(full_path: str):
        """Catch-all: serve React build files or index.html for client-side routing."""
        file_path = static_build_dir / full_path
        if file_path.exists() and file_path.is_file():
            return FileResponse(str(file_path))
        return FileResponse(str(static_build_dir / "index.html"))

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)
